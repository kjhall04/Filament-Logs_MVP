import os
import sqlite3
from contextlib import contextmanager
from datetime import datetime
import json

from backend.config import DATABASE_PATH, DATA_DIR, EXCEL_PATH


def _to_bool(value, default=False):
    if value is None:
        return default
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        return value != 0
    text = str(value).strip().lower()
    if text in ("1", "true", "yes", "on"):
        return True
    if text in ("0", "false", "no", "off"):
        return False
    return default


def _to_int(value, default=0):
    try:
        return int(float(value))
    except (TypeError, ValueError):
        return default


def _to_float(value, default=None):
    try:
        return float(value)
    except (TypeError, ValueError):
        return default


def _normalize_timestamp(value):
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M:%S")
    text = str(value).strip()
    return text or None


_MAPPING_LOOKUP_CACHE = {}
_CANONICALIZATION_SCHEMA_VERSION = 1


def _normalize_space(value):
    return " ".join(str(value or "").split()).strip()


def _lookup_key(value):
    return _normalize_space(value).casefold()


def _load_mapping(filename):
    path = os.path.join(DATA_DIR, filename)
    try:
        with open(path, "r", encoding="utf-8") as handle:
            payload = json.load(handle)
    except Exception:
        return {}
    return payload if isinstance(payload, dict) else {}


def _flatten_color_mapping(color_mapping):
    flattened = {}
    if not isinstance(color_mapping, dict):
        return flattened
    for key_name, value in color_mapping.items():
        if isinstance(value, dict):
            for code, label in value.items():
                flattened[code] = label
        else:
            flattened[key_name] = value
    return flattened


def _mapping_lookup(mapping_name):
    if mapping_name in _MAPPING_LOOKUP_CACHE:
        return _MAPPING_LOOKUP_CACHE[mapping_name]

    if mapping_name == "brand":
        raw = _load_mapping("brand_mapping.json")
    elif mapping_name == "color":
        raw = _flatten_color_mapping(_load_mapping("color_mapping.json"))
    elif mapping_name == "material":
        raw = _load_mapping("material_mapping.json")
    elif mapping_name == "attribute":
        raw = _load_mapping("attribute_mapping.json")
    else:
        raw = {}

    lookup = {}
    for _, label in raw.items():
        canonical = _normalize_space(label)
        if not canonical:
            continue
        lookup[_lookup_key(canonical)] = canonical

    _MAPPING_LOOKUP_CACHE[mapping_name] = lookup
    return lookup


def _canonicalize_mapped_text(value, mapping_name):
    text = _normalize_space(value)
    if not text:
        return ""

    lookup = _mapping_lookup(mapping_name)
    canonical = lookup.get(_lookup_key(text))
    return canonical if canonical is not None else text


def normalize_text_case(value, field=None):
    if value is None:
        return None

    text = _normalize_space(value)
    if not text:
        return ""

    field_key = str(field or "").strip().lower()
    if field_key == "brand":
        return _canonicalize_mapped_text(text, "brand")
    if field_key == "color":
        return _canonicalize_mapped_text(text, "color")
    if field_key == "material":
        return _canonicalize_mapped_text(text, "material")
    if field_key in ("attribute", "attribute_1", "attribute_2"):
        return _canonicalize_mapped_text(text, "attribute")
    if field_key == "location":
        lowered = text.lower()
        if lowered == "lab":
            return "Lab"
        if lowered == "storage":
            return "Storage"
        return text

    return text


def _canonicalize_existing_catalog_values(conn):
    migration_specs = (
        ("inventory", "brand", "brand"),
        ("inventory", "color", "color"),
        ("inventory", "material", "material"),
        ("inventory", "attribute_1", "attribute"),
        ("inventory", "attribute_2", "attribute"),
        ("inventory", "location", "location"),
        ("usage_events", "brand", "brand"),
        ("usage_events", "color", "color"),
        ("usage_events", "material", "material"),
        ("usage_events", "attribute_1", "attribute"),
        ("usage_events", "attribute_2", "attribute"),
        ("usage_events", "location", "location"),
    )

    for table_name, column_name, field_name in migration_specs:
        rows = conn.execute(
            f"""
            SELECT rowid AS __rid, {column_name} AS value
            FROM {table_name}
            WHERE {column_name} IS NOT NULL AND TRIM({column_name}) != ''
            """
        ).fetchall()

        updates = []
        for row in rows:
            current = row["value"]
            canonical = normalize_text_case(current, field=field_name)
            if canonical != current:
                updates.append((canonical, row["__rid"]))

        if updates:
            conn.executemany(
                f"UPDATE {table_name} SET {column_name} = ? WHERE rowid = ?",
                updates,
            )


def _ensure_parent_dir(path):
    parent = os.path.dirname(path) or "."
    os.makedirs(parent, exist_ok=True)


def _load_backup_preferences():
    try:
        from backend import settings_store

        settings = settings_store.load_settings()
    except Exception:
        return False, 30

    enabled = _to_bool(settings.get("auto_backup_on_write"), False)
    retention_days = _to_int(settings.get("backup_retention_days"), 30)
    retention_days = max(1, min(retention_days, 3650))
    return enabled, retention_days


def _cleanup_old_backups(backup_dir, retention_days):
    cutoff = datetime.now().timestamp() - float(retention_days) * 86400.0
    try:
        for name in os.listdir(backup_dir):
            if not name.lower().endswith(".db"):
                continue
            path = os.path.join(backup_dir, name)
            try:
                if os.path.getmtime(path) < cutoff:
                    os.remove(path)
            except Exception:
                continue
    except Exception:
        return


def _backup_database(retention_days):
    if not os.path.exists(DATABASE_PATH):
        return

    database_dir = os.path.dirname(DATABASE_PATH) or "."
    backup_dir = os.path.join(database_dir, "backups")
    os.makedirs(backup_dir, exist_ok=True)

    base_name = os.path.splitext(os.path.basename(DATABASE_PATH))[0]
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup_path = os.path.join(backup_dir, f"{base_name}_{timestamp}.db")

    source = None
    target = None
    try:
        source = sqlite3.connect(DATABASE_PATH)
        target = sqlite3.connect(backup_path)
        source.backup(target)
    except Exception:
        return
    finally:
        if target is not None:
            target.close()
        if source is not None:
            source.close()

    _cleanup_old_backups(backup_dir, retention_days)


def _ensure_schema(conn):
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS inventory (
            timestamp TEXT,
            barcode TEXT PRIMARY KEY,
            brand TEXT,
            color TEXT,
            material TEXT,
            attribute_1 TEXT,
            attribute_2 TEXT,
            filament_amount REAL NOT NULL DEFAULT 0,
            location TEXT,
            roll_weight REAL,
            times_logged_out INTEGER NOT NULL DEFAULT 0,
            is_empty INTEGER NOT NULL DEFAULT 0,
            is_favorite INTEGER NOT NULL DEFAULT 0
        )
        """
    )
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS usage_events (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            timestamp TEXT,
            event_type TEXT,
            barcode TEXT,
            brand TEXT,
            color TEXT,
            material TEXT,
            attribute_1 TEXT,
            attribute_2 TEXT,
            location TEXT,
            input_weight REAL,
            roll_weight REAL,
            filament_amount REAL,
            delta_used REAL,
            times_logged_out INTEGER,
            source TEXT
        )
        """
    )
    conn.execute("CREATE INDEX IF NOT EXISTS idx_inventory_timestamp ON inventory(timestamp)")
    conn.execute(
        "CREATE INDEX IF NOT EXISTS idx_usage_events_type_time ON usage_events(event_type, timestamp)"
    )
    conn.execute("CREATE INDEX IF NOT EXISTS idx_usage_events_barcode ON usage_events(barcode)")

    schema_version = conn.execute("PRAGMA user_version").fetchone()[0]
    if schema_version < _CANONICALIZATION_SCHEMA_VERSION:
        _canonicalize_existing_catalog_values(conn)
        conn.execute(f"PRAGMA user_version = {_CANONICALIZATION_SCHEMA_VERSION}")

    conn.commit()


def _resolve_inventory_sheet(workbook):
    if workbook is None:
        return None
    if "Inventory" in workbook.sheetnames:
        return workbook["Inventory"]
    if workbook.sheetnames:
        return workbook[workbook.sheetnames[0]]
    return None


def _resolve_events_sheet(workbook):
    if workbook is None:
        return None
    if "UsageEvents" in workbook.sheetnames:
        return workbook["UsageEvents"]
    return None


def _import_inventory_rows(conn, inventory_sheet):
    if inventory_sheet is None:
        return 0

    rows = []
    for row in inventory_sheet.iter_rows(min_row=2, values_only=True):
        barcode = ""
        if row and len(row) > 1 and row[1] is not None:
            barcode = str(row[1]).strip()
        if not barcode:
            continue

        rows.append(
            (
                _normalize_timestamp(row[0] if len(row) > 0 else None),
                barcode,
                normalize_text_case(row[2] if len(row) > 2 else None, field="brand"),
                normalize_text_case(row[3] if len(row) > 3 else None, field="color"),
                normalize_text_case(row[4] if len(row) > 4 else None, field="material"),
                normalize_text_case(row[5] if len(row) > 5 else None, field="attribute_1"),
                normalize_text_case(row[6] if len(row) > 6 else None, field="attribute_2"),
                _to_float(row[7] if len(row) > 7 else None, 0.0),
                normalize_text_case(row[8] if len(row) > 8 else None, field="location"),
                _to_float(row[9] if len(row) > 9 else None),
                _to_int(row[10] if len(row) > 10 else None, 0),
                1 if _to_bool(row[11] if len(row) > 11 else None, False) else 0,
                1 if _to_bool(row[12] if len(row) > 12 else None, False) else 0,
            )
        )

    if rows:
        conn.executemany(
            """
            INSERT OR REPLACE INTO inventory (
                timestamp,
                barcode,
                brand,
                color,
                material,
                attribute_1,
                attribute_2,
                filament_amount,
                location,
                roll_weight,
                times_logged_out,
                is_empty,
                is_favorite
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            rows,
        )
    return len(rows)


def _import_event_rows(conn, events_sheet):
    if events_sheet is None:
        return 0

    rows = []
    for row in events_sheet.iter_rows(min_row=2, values_only=True):
        rows.append(
            (
                _normalize_timestamp(row[0] if len(row) > 0 else None),
                row[1] if len(row) > 1 else None,
                str(row[2]).strip() if len(row) > 2 and row[2] is not None else None,
                normalize_text_case(row[3] if len(row) > 3 else None, field="brand"),
                normalize_text_case(row[4] if len(row) > 4 else None, field="color"),
                normalize_text_case(row[5] if len(row) > 5 else None, field="material"),
                normalize_text_case(row[6] if len(row) > 6 else None, field="attribute_1"),
                normalize_text_case(row[7] if len(row) > 7 else None, field="attribute_2"),
                normalize_text_case(row[8] if len(row) > 8 else None, field="location"),
                _to_float(row[9] if len(row) > 9 else None),
                _to_float(row[10] if len(row) > 10 else None),
                _to_float(row[11] if len(row) > 11 else None),
                _to_float(row[12] if len(row) > 12 else None),
                _to_int(row[13] if len(row) > 13 else None, 0),
                row[14] if len(row) > 14 else None,
            )
        )

    if rows:
        conn.executemany(
            """
            INSERT INTO usage_events (
                timestamp,
                event_type,
                barcode,
                brand,
                color,
                material,
                attribute_1,
                attribute_2,
                location,
                input_weight,
                roll_weight,
                filament_amount,
                delta_used,
                times_logged_out,
                source
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            rows,
        )
    return len(rows)


def _try_migrate_from_excel(conn):
    inventory_count = conn.execute("SELECT COUNT(*) FROM inventory").fetchone()[0]
    event_count = conn.execute("SELECT COUNT(*) FROM usage_events").fetchone()[0]
    if inventory_count or event_count:
        return
    if not os.path.exists(EXCEL_PATH):
        return

    try:
        import openpyxl
    except Exception:
        return

    workbook = None
    try:
        workbook = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
        _import_inventory_rows(conn, _resolve_inventory_sheet(workbook))
        _import_event_rows(conn, _resolve_events_sheet(workbook))
        conn.commit()
    except Exception:
        conn.rollback()
    finally:
        if workbook is not None:
            workbook.close()


def convert_excel_to_database(excel_path=None, database_path=None, overwrite=False):
    source_path = os.path.abspath(excel_path or EXCEL_PATH)
    target_path = os.path.abspath(database_path or DATABASE_PATH)

    if not os.path.exists(source_path):
        raise FileNotFoundError(f"Source workbook not found: {source_path}")

    try:
        import openpyxl
    except Exception as exc:
        raise RuntimeError(
            "openpyxl is required for XLSX conversion. Install with: pip install openpyxl"
        ) from exc

    _ensure_parent_dir(target_path)
    if os.path.exists(target_path):
        if not overwrite:
            raise FileExistsError(
                f"Target database already exists: {target_path} (use overwrite=True to replace it)"
            )
        os.remove(target_path)

    conn = sqlite3.connect(target_path, timeout=30)
    conn.row_factory = sqlite3.Row
    workbook = None
    try:
        _ensure_schema(conn)

        workbook = openpyxl.load_workbook(source_path, data_only=True)
        inventory_rows = _import_inventory_rows(conn, _resolve_inventory_sheet(workbook))
        event_rows = _import_event_rows(conn, _resolve_events_sheet(workbook))
        conn.commit()

        return {
            "excel_path": source_path,
            "database_path": target_path,
            "inventory_rows": inventory_rows,
            "event_rows": event_rows,
        }
    except Exception:
        conn.rollback()
        raise
    finally:
        if workbook is not None:
            workbook.close()
        conn.close()


@contextmanager
def open_database(write=False):
    _ensure_parent_dir(DATABASE_PATH)

    if write:
        backup_enabled, backup_retention_days = _load_backup_preferences()
        if backup_enabled:
            _backup_database(backup_retention_days)

    conn = sqlite3.connect(DATABASE_PATH, timeout=30)
    conn.row_factory = sqlite3.Row

    try:
        _ensure_schema(conn)
        _try_migrate_from_excel(conn)
        yield conn
        if write:
            conn.commit()
    except Exception:
        if write:
            conn.rollback()
        raise
    finally:
        conn.close()


def _inventory_row_to_tuple(row):
    return (
        row["timestamp"],
        row["barcode"],
        normalize_text_case(row["brand"], field="brand"),
        normalize_text_case(row["color"], field="color"),
        normalize_text_case(row["material"], field="material"),
        normalize_text_case(row["attribute_1"], field="attribute_1"),
        normalize_text_case(row["attribute_2"], field="attribute_2"),
        _to_float(row["filament_amount"], 0.0),
        normalize_text_case(row["location"], field="location"),
        _to_float(row["roll_weight"]),
        _to_int(row["times_logged_out"], 0),
        "True" if _to_bool(row["is_empty"], False) else "False",
        "True" if _to_bool(row["is_favorite"], False) else "False",
    )


def _inventory_row_to_dict(row):
    if row is None:
        return None

    return {
        "timestamp": row["timestamp"],
        "barcode": str(row["barcode"]).strip() if row["barcode"] is not None else "",
        "brand": normalize_text_case(row["brand"], field="brand"),
        "color": normalize_text_case(row["color"], field="color"),
        "material": normalize_text_case(row["material"], field="material"),
        "attribute_1": normalize_text_case(row["attribute_1"], field="attribute_1"),
        "attribute_2": normalize_text_case(row["attribute_2"], field="attribute_2"),
        "filament_amount": _to_float(row["filament_amount"], 0.0),
        "location": normalize_text_case(row["location"], field="location"),
        "roll_weight": _to_float(row["roll_weight"]),
        "times_logged_out": _to_int(row["times_logged_out"], 0),
        "is_empty": _to_bool(row["is_empty"], False),
        "is_favorite": _to_bool(row["is_favorite"], False),
    }


def list_inventory_rows():
    with open_database(write=False) as conn:
        rows = conn.execute(
            """
            SELECT
                timestamp,
                barcode,
                brand,
                color,
                material,
                attribute_1,
                attribute_2,
                filament_amount,
                location,
                roll_weight,
                times_logged_out,
                is_empty,
                is_favorite
            FROM inventory
            ORDER BY rowid ASC
            """
        ).fetchall()
        return [_inventory_row_to_tuple(row) for row in rows]


def list_inventory_barcodes(conn=None):
    if conn is not None:
        rows = conn.execute(
            "SELECT barcode FROM inventory WHERE barcode IS NOT NULL AND TRIM(barcode) != ''"
        ).fetchall()
    else:
        with open_database(write=False) as read_conn:
            rows = read_conn.execute(
                "SELECT barcode FROM inventory WHERE barcode IS NOT NULL AND TRIM(barcode) != ''"
            ).fetchall()
    return [str(row["barcode"]).strip() for row in rows if row["barcode"] is not None]


def get_roll_weight(barcode: str, conn=None):
    if not barcode:
        return None

    target = str(barcode).strip()
    if not target:
        return None

    def _resolve_from_row(row):
        if row is None:
            return None
        for key in ("roll_weight", "filament_amount", "times_logged_out"):
            value = _to_float(row[key] if key in row.keys() else None)
            if value is not None:
                return value
        return None

    if conn is not None:
        row = conn.execute(
            """
            SELECT roll_weight, filament_amount, times_logged_out
            FROM inventory
            WHERE barcode = ?
            LIMIT 1
            """,
            (target,),
        ).fetchone()
        return _resolve_from_row(row)

    with open_database(write=False) as read_conn:
        row = read_conn.execute(
            """
            SELECT roll_weight, filament_amount, times_logged_out
            FROM inventory
            WHERE barcode = ?
            LIMIT 1
            """,
            (target,),
        ).fetchone()
        return _resolve_from_row(row)


def toggle_inventory_favorite(barcode: str):
    if not barcode:
        return None

    target = str(barcode).strip()
    if not target:
        return None

    with open_database(write=True) as conn:
        cursor = conn.execute(
            """
            UPDATE inventory
            SET is_favorite = CASE WHEN COALESCE(is_favorite, 0) = 1 THEN 0 ELSE 1 END
            WHERE barcode = ?
            """,
            (target,),
        )
        if cursor.rowcount <= 0:
            return None

        row = conn.execute(
            "SELECT is_favorite FROM inventory WHERE barcode = ? LIMIT 1",
            (target,),
        ).fetchone()
        if row is None:
            return None
        return _to_bool(row["is_favorite"], False)


def get_inventory_roll(barcode: str, conn=None):
    if not barcode:
        return None

    target = str(barcode).strip()
    if not target:
        return None

    query = """
        SELECT
            timestamp,
            barcode,
            brand,
            color,
            material,
            attribute_1,
            attribute_2,
            filament_amount,
            location,
            roll_weight,
            times_logged_out,
            is_empty,
            is_favorite
        FROM inventory
        WHERE barcode = ?
        LIMIT 1
    """

    if conn is not None:
        row = conn.execute(query, (target,)).fetchone()
        return _inventory_row_to_dict(row)

    with open_database(write=False) as read_conn:
        row = read_conn.execute(query, (target,)).fetchone()
        return _inventory_row_to_dict(row)


def update_inventory_roll(
    barcode: str,
    brand,
    color,
    material,
    attribute_1,
    attribute_2,
    location,
    filament_amount,
    roll_weight,
    is_empty,
):
    if not barcode:
        return False

    target = str(barcode).strip()
    if not target:
        return False

    brand_value = normalize_text_case(brand, field="brand")
    color_value = normalize_text_case(color, field="color")
    material_value = normalize_text_case(material, field="material")
    attribute_1_value = normalize_text_case(attribute_1, field="attribute_1")
    attribute_2_value = normalize_text_case(attribute_2, field="attribute_2")
    location_value = normalize_text_case(location, field="location")
    filament_amount_value = round(max(_to_float(filament_amount, 0.0), 0.0), 2)
    roll_weight_value = _to_float(roll_weight)
    is_empty_value = 1 if _to_bool(is_empty, False) else 0

    with open_database(write=True) as conn:
        cursor = conn.execute(
            """
            UPDATE inventory
            SET
                brand = ?,
                color = ?,
                material = ?,
                attribute_1 = ?,
                attribute_2 = ?,
                location = ?,
                filament_amount = ?,
                roll_weight = ?,
                is_empty = ?
            WHERE barcode = ?
            """,
            (
                brand_value,
                color_value,
                material_value,
                attribute_1_value,
                attribute_2_value,
                location_value,
                filament_amount_value,
                roll_weight_value,
                is_empty_value,
                target,
            ),
        )
        if cursor.rowcount <= 0:
            return False

        conn.execute(
            """
            UPDATE usage_events
            SET
                brand = ?,
                color = ?,
                material = ?,
                attribute_1 = ?,
                attribute_2 = ?,
                location = ?
            WHERE barcode = ?
            """,
            (
                brand_value,
                color_value,
                material_value,
                attribute_1_value,
                attribute_2_value,
                location_value,
                target,
            ),
        )

    return True
